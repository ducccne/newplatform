import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# BẢNG TỪ ĐỒNG NGHĨA CHỮ 
COLUMN_MAPPING = {
    'Warning message ID': ['Warning message ID', 'Warning_Msg_ID', 'Msg ID', 'Warning ID', 'Mã cảnh báo'],
    'Warning title - EN (New)': ['Warning title - EN (New)', 'Warning title', 'Warning Title-EN'],
    'Warning content - EN (New)': ['Warning content - EN (New)', 'Warning content', 'Warning Content-EN'],
    'Full Warning Content - EN (New)': ['Full Warning Content - EN (New)', 'Full Warning Content-EN'],
    'Warning message icon': ['Warning message icon', 'Icon', 'Warning icon', 'Icon ID'],
    'Warning acoustic signal (New)': ['Warning acoustic signal (New)', 'Warning acoustic signal', 'Acoustic', 'Sound'],
    '__COL_COND_SET__': [
        'Conditions to set warning (Reference)', 'Conditions to set warning', 
        'condition to set', 'condition_to_set', 'set_condition', 'to set'
    ],
    '__COL_COND_SHOW__': [
        'Conditions to show warning', 'conditions to show warning', 
        'condition to show', 'condition_to_show', 'show_condition', 'to show'
    ],
    
    'Set requirement for Head Unit': ['Set requirement for Head Unit', 'Set_Requirement', 'Requirement', 'Condition', 'Yêu cầu kích hoạt']
}

def detect_excel_layout(ws):
    """
    Quét nhanh file Excel để xác định tọa độ dòng tiêu đề và vị trí cột ID 
    """
    detected_head_row = None
    detected_id_col = None

    # Quét radar 10 dòng đầu tiên để dò tìm ô chứa từ khóa 'ID'
    for r in range(1, 11):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(r, c).value).strip().lower()
            if val == "id":
                detected_head_row = r
                detected_id_col = c
                break
        if detected_head_row: 
            break

    if not detected_head_row:
        print("Không tự động tìm thấy dòng Header")
        return 2, 3, 1

    detected_data_start_row = detected_head_row + 1
    
    print(f"Phân tích cấu trúc File:")
    print(f"Dòng chứa tiêu đề (Header Row): {detected_head_row}")
    print(f"Dòng bắt đầu dữ liệu (Data Start): {detected_data_start_row}")
    print(f"Vị trí cột ID: {detected_id_col}")
    
    return detected_head_row, detected_data_start_row, detected_id_col

def check_if_excel_has_merged_cells(file_path):
    """Kiểm tra cấu trúc ô: Trả về True nếu phát hiện file chứa ô bị trộn gộp (Merge)."""
    try:
        wb = load_workbook(file_path, read_only=True)
        return len(wb.active.merged_cells.ranges) > 0
    except Exception:
        return False

# file kiểu 2: có 2 cột conditions to set/show warning
def standardize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Chuẩn hóa tên cột đồng nghĩa.
    Xử lý logic File kiểu 2: Ưu tiên nhặt câu lệnh cột Set, nếu ô bị trống tự động bù bằng ô cột Show.
    """
    current_to_std = {}
    for current_col in df.columns:
        current_col_clean = str(current_col).strip().lower()
        for std_name, synonyms in COLUMN_MAPPING.items():
            if current_col_clean in [str(s).strip().lower() for s in synonyms]:
                current_to_std[current_col] = std_name
                break
                
    df = df.rename(columns=current_to_std)
    
    has_cond_set = '__COL_COND_SET__' in df.columns
    has_cond_show = '__COL_COND_SHOW__' in df.columns
    has_set_req = '__COL_SET_REQ__' in df.columns
    
    # ─── THỰC THI THUẬT TOÁN ĐỘ ƯU TIÊN CỘT ĐIỀU KIỆN ───
    
    if has_cond_set or has_cond_show:
        print("--> [Column Resolver]: Thực thi logic FILE KIỂU 2. Đang gộp ưu tiên cột Set -> Show...")
        
        # [SỬA LỖI TẠI ĐÂY] Ép trước các chuỗi rỗng, khoảng trắng thành NaN để combine_first() nhận diện được
        if has_cond_set:
            df['__COL_COND_SET__'] = df['__COL_COND_SET__'].replace([r'^\s*$', 'nan', 'None'], np.nan, regex=True)
        if has_cond_show:
            df['__COL_COND_SHOW__'] = df['__COL_COND_SHOW__'].replace([r'^\s*$', 'nan', 'None'], np.nan, regex=True)
            
        df['Set requirement for Head Unit'] = np.nan
        
        # Bước 1: Đổ dữ liệu cột Show làm nền trước (Ưu tiên thấp hơn)
        if has_cond_show:
            df['Set requirement for Head Unit'] = df['Set requirement for Head Unit'].fillna(df['__COL_COND_SHOW__'])
            
        # Bước 2: Lấy dữ liệu cột Set ghi đè lên. Ô nào cột Set là NaN thì giữ nguyên nền cột Show.
        if has_cond_set:
            df['Set requirement for Head Unit'] = df['__COL_COND_SET__'].combine_first(df['Set requirement for Head Unit'])
            
    # TÌNH HUỐNG B: FILE KIỂU 1 (Mẫu truyền thống có sẵn cột gộp chung điều kiện)
    elif has_set_req:
        print("--> [Column Resolver]: Thực thi logic FILE KIỂU 1. Lấy thẳng dữ liệu cột 'Set requirement for Head Unit'.")
        df['Set requirement for Head Unit'] = df['__COL_SET_REQ__']
        
    else:
        print("Không tìm thấy bất kỳ cột điều kiện nào đạt chuẩn!")
        df['Set requirement for Head Unit'] = np.nan

    # Giải phóng và xóa bỏ các cột nhãn tạm thời
    cols_to_drop = [c for c in ['__COL_COND_SET__', '__COL_COND_SHOW__', '__COL_SET_REQ__'] if c in df.columns]
    df = df.drop(columns=cols_to_drop)
    
    return df

def clean_wiki_and_rich_text(text):
    if not isinstance(text, str): return text
    text = re.sub(r'%%\((?:[^()]+|\([^()]*\))*\)', '', text)
    text = re.sub(r'\[\{Image wiki=[\'"](.*?)[\'"].*?\}\]', r'\1', text)
    text = text.replace('%!', '')
    text = text.replace('_x000D_x000D_', '\n').replace('_x000D_', '\n')
    text = re.sub(r'~([_A-Za-z0-9])', r'\1', text)
    text = text.replace('\r', '')
    text = re.sub(r'\n\s*\n', '\n', text)
    return text.strip()

def load_excel_and_clean(file_path, mapping_icon_path="mapping_icon.xlsx"):
    """Pipeline kết nối: Tự động dò layout, gỡ gộp đồng bộ mọi cột, xử lý ép chuẩn cột điều kiện."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Thu thập tọa độ động từ radar quét
    head_row, data_start_row, id_column = detect_excel_layout(ws)
    has_merged = check_if_excel_has_merged_cells(file_path)
    
    if has_merged:
        print("--> [Data Processor]: Cấu trúc: FILE CÓ MERGE CELL. Đang tiến hành duỗi phẳng đồng bộ...")
        
        # 1. Tạo bản đồ tra cứu giá trị Master cho toàn bộ các khối ô gộp (Áp dụng đồng bộ mọi cột: ID, Title,...)
        slave_master = {}
        for mc in ws.merged_cells.ranges:
            master_val = ws.cell(mc.min_row, mc.min_col).value
            for row in range(mc.min_row, mc.max_row + 1):
                for col in range(mc.min_col, mc.max_col + 1):
                    slave_master[(row, col)] = master_val

        # 2. Xây dựng cây chỉ mục liên kết Hyperlink từ cột ID gốc
        id_merge_group = {}
        for mc in ws.merged_cells.ranges:
            if mc.min_col == id_column and mc.min_row >= data_start_row and mc.max_row > mc.min_row:
                for row in range(mc.min_row, mc.max_row + 1):
                    id_merge_group[row] = (mc.min_row, mc.max_row)

        id_links = {}
        for r in range(data_start_row, ws.max_row + 1):
            m_row = id_merge_group[r][0] if r in id_merge_group else r
            cell = ws.cell(m_row, id_column)
            id_links[r] = cell.hyperlink.target if cell.hyperlink else ""

        # 3. Đọc danh sách tên cột tiêu đề
        raw_headers = [ws.cell(head_row, col).value for col in range(1, ws.max_column + 1)]
        last_valid_col = max(i + 1 for i, h in enumerate(raw_headers) if h is not None)
        headers = raw_headers[:last_valid_col]

        # 4. Trích xuất dữ liệu: Ô nào gộp tự động lấy Master fill xuống, ô nào phẳng giữ nguyên vẹn độc lập
        final_rows = []
        for row_num in range(data_start_row, ws.max_row + 1):
            row_data = []
            for col_num in range(1, last_valid_col + 1):
                if (row_num, col_num) in slave_master:
                    val = slave_master[(row_num, col_num)]
                else:
                    val = ws.cell(row_num, col_num).value
                row_data.append(val)
            
            # Gắn kèm link gốc của Requirement vào cuối hàng phục vụ rebuild hyperlink đầu ra
            row_data.append(id_links.get(row_num, ""))
            final_rows.append(row_data)

        df = pd.DataFrame(final_rows, columns=list(headers)[:last_valid_col] + ["_ID_Hyperlink_Target"])
    else:
        print("--> [Data Processor]: Cấu trúc: FILE PHẲNG HOÀN TOÀN. Kích hoạt Pandas đọc trực tiếp...")
        df_raw = pd.read_excel(file_path, header=head_row - 1)
        
        id_links = {r - (head_row + 1): ws.cell(r, id_column).hyperlink.target if ws.cell(r, id_column).hyperlink else "" for r in range(data_start_row, ws.max_row + 1)}
        df_raw["_ID_Hyperlink_Target"] = df_raw.index.map(id_links).fillna("")
        df = df_raw

    # Chuẩn hóa rẽ nhánh và chọn lọc cột điều kiện kỹ thuật
    df = standardize_dataframe_columns(df)
    
    print("--> [Data Processor]: Đang dọn dẹp làm sạch toàn bộ ký tự rác hệ thống (Wiki Markup & CSS)...")
    for col in df.columns:
        if col != "_ID_Hyperlink_Target": 
            df[col] = df[col].apply(clean_wiki_and_rich_text)
            
    df.replace(['nan', '', ' '], np.nan, inplace=True)
    return df

def extract_clean_image_name(text):
    if pd.isna(text) or not isinstance(text, str): return ""
    match = re.search(r'([\w\-_]+\.png)', text, re.IGNORECASE)
    return match.group(1).strip() if match else text.strip()

def load_icon_dictionary(mapping_file):
    if not os.path.exists(mapping_file): return {}
    try:
        df_map = pd.read_excel(mapping_file)
        col_image, col_id = df_map.columns[0], df_map.columns[2] if len(df_map.columns) >= 3 else df_map.columns[1]
        df_map['Clean_Key'] = df_map[col_image].apply(extract_clean_image_name)
        df_map = df_map.dropna(subset=['Clean_Key', col_id])
        return dict(zip(df_map['Clean_Key'], df_map[col_id]))
    except Exception: return {}

def apply_smart_icon_mapping(df: pd.DataFrame, icon_dict: dict):
    if "Warning message icon" not in df.columns or not icon_dict: return df
    df['temp_row_idx'] = df.groupby("ID Requirement").cumcount()
    def map_correct_icon(row):
        raw_text, row_idx = str(row["Warning message icon"]), int(row['temp_row_idx'])
        if raw_text.strip() in ['nan', 'null', '', 'None']: return np.nan
        all_images = re.findall(r'([\w\-_]+\.(?:png|jpg|jpeg))', raw_text, re.IGNORECASE)
        if not all_images: return raw_text
        target_image = all_images[row_idx] if row_idx < len(all_images) else all_images[-1]
        return icon_dict.get(target_image, target_image)
    df["Warning message icon"] = df.apply(map_correct_icon, axis=1)
    return df.drop(columns=['temp_row_idx'])

def build_testcase_structure(df_clean, mapping_icon_path="mapping_icon.xlsx"):
    testcases = []
    root_folder_id, root_folder_summary, immediate_parent_folder = "", "", ""

    for index, row in df_clean.iterrows():
        row_type, row_id, summary = str(row.get('Type', '')).strip(), str(row.get('ID', '')).strip(), str(row.get('Summary', '')).strip()
        if row_id.endswith('.0'): row_id = row_id[:-2]
        if root_folder_id.endswith('.0'): root_folder_id = root_folder_id[:-2]

        if row_type == "Folder":
            if pd.notna(row_id) and row_id != "" and row_id != "nan": root_folder_id = row_id
            root_folder_summary = immediate_parent_folder = summary
            continue
        if pd.isna(row_id) or row_id == "" or row_id == "nan": continue

        tc_record = row.to_dict()
        tc_record["Task_ID"] = f"Task_{index}"
        tc_record["ID Requirement"] = row_id
        tc_record["Functional Level 1"] = immediate_parent_folder
        tc_record["FIP_ID"] = root_folder_id
        tc_record["FIP_ID & Summary"] = root_folder_summary
        tc_record["FIF_Summary"] = summary
        tc_record["Priority"] = str(row.get('Category', '')).replace('nan', '')
        tc_record["Environment"], tc_record["Vehicle Status"] = "All", "Static"
        tc_record["Power State"] = "ACC, On Driver Present, ON Emotor Running"
        tc_record["Precondition signal"], tc_record["Test Steps.signal input CAN/Lin signal"] = "N/A", "N/A"
        tc_record["Variant"], tc_record["Market"], tc_record["Program"] = "N/A", "N/A", "All"
        tc_record["Name"] = summary
        tc_record["Test Purpose"] = tc_record["Pre-Action"] = tc_record["Test Steps.Action"] = tc_record["Test Steps.Expected result"] = "" 
        testcases.append(tc_record)
        
    df_output = pd.DataFrame(testcases)
    icon_dict = load_icon_dictionary(mapping_icon_path)
    if icon_dict and "Warning message icon" in df_output.columns:
        df_output = apply_smart_icon_mapping(df_output, icon_dict)
    return df_output